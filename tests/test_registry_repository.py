import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from file_compare_app.registry.database import connect_registry_db, default_registry_db_path, initialize_registry_db
from file_compare_app.registry.repository import RegistryRepository


def test_default_registry_database_is_next_to_frozen_executable(tmp_path: Path, monkeypatch) -> None:
    executable = tmp_path / "dist" / "file-compare.exe"
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(executable))

    assert default_registry_db_path() == executable.parent / "registry.sqlite3"


def test_connect_registry_db_copies_legacy_database_next_to_executable(tmp_path: Path, monkeypatch) -> None:
    legacy_root = tmp_path / "localappdata"
    legacy_db = legacy_root / "FileCompareRegistry" / "registry.sqlite3"
    executable = tmp_path / "dist" / "file-compare.exe"
    executable.parent.mkdir()
    monkeypatch.setenv("LOCALAPPDATA", str(legacy_root))
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(executable))

    repository = RegistryRepository.create(legacy_db)
    repository.create_node(name="Legacy root")
    repository.connection.close()

    with connect_registry_db() as connection:
        rows = connection.execute("select name from registry_nodes").fetchall()

    assert default_registry_db_path().exists()
    assert [row["name"] for row in rows] == ["Legacy root"]


def test_registry_database_creates_schema(tmp_path: Path) -> None:
    db_path = tmp_path / "registry.sqlite3"

    with connect_registry_db(db_path) as connection:
        initialize_registry_db(connection)
        table_names = {
            row[0]
            for row in connection.execute(
                "select name from sqlite_master where type = 'table'"
            ).fetchall()
        }

    assert {"registry_nodes", "watched_files", "file_versions", "monitor_events"}.issubset(table_names)


def test_repository_creates_root_and_nested_nodes(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")

    root = repository.create_node(name="Завод №1", node_type="Объект", code="OBJ-001")
    child = repository.create_node(name="Цех 1", parent_id=root.id, node_type="Подобъект")
    grandchild = repository.create_node(name="Линия упаковки", parent_id=child.id)

    assert root.parent_id is None
    assert child.parent_id == root.id
    assert grandchild.parent_id == child.id
    assert [node.name for node in repository.list_children(root.id)] == ["Цех 1"]
    assert [node.name for node in repository.list_children(child.id)] == ["Линия упаковки"]


def test_repository_creates_watched_file_and_event(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    node = repository.create_node(name="Линия упаковки")

    watched = repository.create_watched_file(
        node_id=node.id,
        file_path=r"D:\Objects\Factory-1\controller.yaml",
        label="controller.yaml",
        file_kind="config",
        baseline_hash="abc",
        last_hash="abc",
    )
    event = repository.create_event(
        node_id=node.id,
        watched_file_id=watched.id,
        event_type="baseline",
        message="baseline saved",
    )

    assert watched.node_id == node.id
    assert watched.status == "ok"
    assert repository.list_watched_files(node.id)[0].label == "controller.yaml"
    assert repository.list_events(node.id)[0].id == event.id


def test_repository_stores_ordered_file_versions(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    node = repository.create_node(name="Kaspersky")
    watched = repository.create_watched_file(
        node_id=node.id,
        file_path=str(tmp_path / "policy.klp"),
        label="policy.klp",
        file_kind="klp",
        baseline_hash="old",
        last_hash="old",
    )

    first = repository.create_file_version(
        watched_file_id=watched.id,
        version_number=1,
        snapshot_path=str(tmp_path / "snapshots" / "v0001_policy.klp"),
        sha256="old",
        size_bytes=100,
    )
    second = repository.create_file_version(
        watched_file_id=watched.id,
        version_number=2,
        snapshot_path=str(tmp_path / "snapshots" / "v0002_policy.klp"),
        sha256="new",
        size_bytes=120,
        change_count=3,
    )

    assert repository.get_latest_file_version(watched.id).id == second.id
    assert [version.id for version in repository.list_file_versions(watched.id)] == [first.id, second.id]


def test_repository_deletes_node_with_children_files_and_events(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    root = repository.create_node(name="Root")
    child = repository.create_node(name="Child", parent_id=root.id)
    grandchild = repository.create_node(name="Grandchild", parent_id=child.id)
    watched = repository.create_watched_file(
        node_id=grandchild.id,
        file_path=r"D:\Objects\Factory-1\controller.yaml",
        label="controller.yaml",
        file_kind="config",
        baseline_hash="abc",
        last_hash="abc",
    )
    event = repository.create_event(
        node_id=grandchild.id,
        watched_file_id=watched.id,
        event_type="baseline",
        message="baseline saved",
    )

    deleted_ids = repository.delete_node(child.id)

    assert set(deleted_ids) == {child.id, grandchild.id}
    assert repository.list_children(root.id) == []
    with pytest.raises(KeyError):
        repository.get_node(child.id)
    with pytest.raises(KeyError):
        repository.get_node(grandchild.id)
    with pytest.raises(KeyError):
        repository.get_watched_file(watched.id)
    with pytest.raises(KeyError):
        repository.get_event(event.id)


def test_repository_updates_registry_node_card_fields(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    node = repository.create_node(name="Object", code="OLD", responsible="Old owner")

    updated = repository.update_node(
        node.id,
        name="Main object",
        code="OBJ-2026",
        responsible="Engineer",
        description="Critical control node",
    )

    assert updated.name == "Main object"
    assert updated.code == "OBJ-2026"
    assert updated.responsible == "Engineer"
    assert updated.description == "Critical control node"
    assert repository.get_node(node.id).updated_at


def test_repository_creates_database_backup_next_to_database(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    repository.create_node(name="Backed up object")

    backup_path = repository.backup_database()

    assert backup_path.parent == tmp_path / "backups"
    assert backup_path.name.startswith("registry-backup-")
    assert backup_path.suffix == ".sqlite3"
    assert backup_path.exists()

    restored = RegistryRepository.create(backup_path)
    assert restored.list_children(None)[0].name == "Backed up object"


def test_repository_exports_registry_to_excel(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    node = repository.create_node(
        name="Factory",
        code="OBJ-001",
        responsible="Operator",
        description="Production line",
    )
    watched = repository.create_watched_file(
        node_id=node.id,
        file_path=str(tmp_path / "running-config.cfg"),
        label="running-config.cfg",
        file_kind="cfg",
        baseline_hash="old",
        last_hash="new",
        status="changed",
    )
    repository.create_file_version(
        watched_file_id=watched.id,
        version_number=1,
        snapshot_path=str(tmp_path / "v0001_running-config.cfg"),
        sha256="old",
        size_bytes=10,
    )
    repository.create_file_version(
        watched_file_id=watched.id,
        version_number=2,
        snapshot_path=str(tmp_path / "v0002_running-config.cfg"),
        sha256="new",
        size_bytes=12,
        change_count=2,
    )
    repository.create_event(
        node_id=node.id,
        watched_file_id=watched.id,
        event_type="changed",
        change_count=2,
        message="v1 -> v2",
    )
    repository.upsert_comparison_review(
        watched_file_id=watched.id,
        from_version=1,
        to_version=2,
        status="approved",
        reviewer="Amal",
        comment="Проверено",
    )

    export_path = repository.export_registry_excel(tmp_path / "registry.xlsx")
    workbook = load_workbook(export_path)
    compare_sheet = workbook["Сверки"]

    assert workbook.sheetnames == ["Объекты", "Файлы", "Сверки", "История"]
    assert workbook["Объекты"]["D2"].value == "Factory"
    assert workbook["Объекты"]["E2"].value == "OBJ-001"
    assert workbook["Файлы"]["D2"].value == "running-config.cfg"
    assert workbook["Файлы"]["F2"].value == "Изменен"
    assert compare_sheet["E2"].value == "v1 -> v2"
    assert compare_sheet["H2"].value == 2
    assert compare_sheet["J1"].value == "Согласование"
    assert compare_sheet["J2"].value == "Согласовано"
    assert compare_sheet["K2"].value == "Amal"
    assert compare_sheet["L2"].value == "Проверено"
    assert workbook["История"]["E2"].value == "Изменен"


def test_repository_stores_node_template_and_monitoring_schedule(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")

    node = repository.create_node(
        name="Kaspersky ARM",
        template_key="kaspersky",
        monitor_interval_minutes=120,
    )
    updated = repository.update_node(
        node.id,
        template_key="cisco_asa",
        monitoring_enabled=False,
        monitor_interval_minutes=240,
    )

    assert node.template_key == "kaspersky"
    assert node.monitor_interval_minutes == 120
    assert updated.template_key == "cisco_asa"
    assert updated.monitoring_enabled is False
    assert updated.monitor_interval_minutes == 240


def test_repository_upserts_comparison_review_for_version_pair(tmp_path: Path) -> None:
    repository = RegistryRepository.create(tmp_path / "registry.sqlite3")
    node = repository.create_node(name="Object")
    watched = repository.create_watched_file(
        node_id=node.id,
        file_path=str(tmp_path / "running.cfg"),
        label="running.cfg",
        file_kind="cfg",
        baseline_hash="old",
        last_hash="new",
        status="changed",
    )

    review = repository.upsert_comparison_review(
        watched_file_id=watched.id,
        from_version=1,
        to_version=2,
        status="approved",
        reviewer="Amal",
        comment="Согласовано после проверки",
    )
    updated = repository.upsert_comparison_review(
        watched_file_id=watched.id,
        from_version=1,
        to_version=2,
        status="rejected",
        reviewer="Engineer",
        comment="Нужно уточнение",
    )

    assert review.status == "approved"
    assert updated.id == review.id
    assert updated.status == "rejected"
    assert updated.reviewer == "Engineer"
    assert repository.get_comparison_review(watched.id, 1, 2).comment == "Нужно уточнение"
