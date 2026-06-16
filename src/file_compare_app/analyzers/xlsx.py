from pathlib import Path

from file_compare_app.analyzers.base import Analyzer, AnalyzerDependencyError
from file_compare_app.core.models import Change, ChangeLocation, ComparisonResult


class XlsxAnalyzer(Analyzer):
    def compare(self, left_path: Path, right_path: Path) -> ComparisonResult:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise AnalyzerDependencyError("Для сравнения Excel установите openpyxl.") from exc

        left_wb = load_workbook(left_path, data_only=False)
        right_wb = load_workbook(right_path, data_only=False)
        changes: list[Change] = []
        index = 1
        for sheet_name in sorted(set(left_wb.sheetnames) | set(right_wb.sheetnames)):
            if sheet_name not in left_wb.sheetnames:
                changes.append(_sheet_change(index, left_path, right_path, sheet_name, "", "sheet added", "added"))
                index += 1
                continue
            if sheet_name not in right_wb.sheetnames:
                changes.append(_sheet_change(index, left_path, right_path, sheet_name, "sheet removed", "", "removed"))
                index += 1
                continue
            left_ws = left_wb[sheet_name]
            right_ws = right_wb[sheet_name]
            max_row = max(left_ws.max_row, right_ws.max_row)
            max_col = max(left_ws.max_column, right_ws.max_column)
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    left_cell = left_ws.cell(row=row, column=col)
                    right_cell = right_ws.cell(row=row, column=col)
                    if left_cell.value == right_cell.value:
                        continue
                    changes.append(
                        Change(
                            change_id=f"xlsx-{index}",
                            change_type="modified",
                            severity="normal",
                            source_location=ChangeLocation(str(left_path), sheet=sheet_name, cell=left_cell.coordinate),
                            target_location=ChangeLocation(str(right_path), sheet=sheet_name, cell=right_cell.coordinate),
                            before="" if left_cell.value is None else str(left_cell.value),
                            after="" if right_cell.value is None else str(right_cell.value),
                            format="xlsx",
                            confidence=1.0,
                            notes=f"{sheet_name}!{right_cell.coordinate}",
                        )
                    )
                    index += 1
        return ComparisonResult(str(left_path), str(right_path), "xlsx", changes)


def _sheet_change(index: int, left_path: Path, right_path: Path, sheet: str, before: str, after: str, change_type: str) -> Change:
    return Change(
        change_id=f"xlsx-{index}",
        change_type=change_type,  # type: ignore[arg-type]
        severity="normal",
        source_location=ChangeLocation(str(left_path), sheet=sheet),
        target_location=ChangeLocation(str(right_path), sheet=sheet),
        before=before,
        after=after,
        format="xlsx",
        confidence=1.0,
        notes=sheet,
    )
