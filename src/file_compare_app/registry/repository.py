from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from file_compare_app.registry.database import connect_registry_db, initialize_registry_db
from file_compare_app.registry.models import ComparisonReview, FileVersion, MonitorEvent, RegistryNode, WatchedFile


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


class RegistryRepository:
    def __init__(self, connection: sqlite3.Connection) -> None:
        self.connection = connection
        self.db_path = _connection_path(connection)

    @classmethod
    def create(cls, path: Path | None = None) -> "RegistryRepository":
        connection = connect_registry_db(path)
        initialize_registry_db(connection)
        return cls(connection)

    def create_node(
        self,
        name: str,
        parent_id: int | None = None,
        node_type: str = "Объект",
        code: str = "",
        responsible: str = "",
        description: str = "",
        template_key: str = "",
        monitoring_enabled: bool = True,
        monitor_interval_minutes: int = 30,
    ) -> RegistryNode:
        timestamp = _now()
        cursor = self.connection.execute(
            """
            insert into registry_nodes
            (parent_id, name, node_type, code, responsible, description, template_key, monitoring_enabled, monitor_interval_minutes, created_at, updated_at)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                parent_id,
                name,
                node_type,
                code,
                responsible,
                description,
                template_key,
                1 if monitoring_enabled else 0,
                monitor_interval_minutes,
                timestamp,
                timestamp,
            ),
        )
        self.connection.commit()
        return self.get_node(int(cursor.lastrowid))

    def get_node(self, node_id: int) -> RegistryNode:
        row = self.connection.execute(
            "select * from registry_nodes where id = ?",
            (node_id,),
        ).fetchone()
        if row is None:
            raise KeyError(f"registry node {node_id} not found")
        return _node_from_row(row)

    def update_node(
        self,
        node_id: int,
        *,
        name: str | None = None,
        node_type: str | None = None,
        code: str | None = None,
        responsible: str | None = None,
        description: str | None = None,
        template_key: str | None = None,
        monitoring_enabled: bool | None = None,
        monitor_interval_minutes: int | None = None,
    ) -> RegistryNode:
        current = self.get_node(node_id)
        timestamp = _now()
        self.connection.execute(
            """
            update registry_nodes
            set name = ?,
                node_type = ?,
                code = ?,
                responsible = ?,
                description = ?,
                template_key = ?,
                monitoring_enabled = ?,
                monitor_interval_minutes = ?,
                updated_at = ?
            where id = ?
            """,
            (
                current.name if name is None else name,
                current.node_type if node_type is None else node_type,
                current.code if code is None else code,
                current.responsible if responsible is None else responsible,
                current.description if description is None else description,
                current.template_key if template_key is None else template_key,
                int(current.monitoring_enabled if monitoring_enabled is None else monitoring_enabled),
                current.monitor_interval_minutes if monitor_interval_minutes is None else monitor_interval_minutes,
                timestamp,
                node_id,
            ),
        )
        self.connection.commit()
        return self.get_node(node_id)

    def list_children(self, parent_id: int | None) -> list[RegistryNode]:
        if parent_id is None:
            rows = self.connection.execute(
                "select * from registry_nodes where parent_id is null order by name"
            ).fetchall()
        else:
            rows = self.connection.execute(
                "select * from registry_nodes where parent_id = ? order by name",
                (parent_id,),
            ).fetchall()
        return [_node_from_row(row) for row in rows]

    def list_descendant_ids(self, node_id: int) -> list[int]:
        rows = self.connection.execute(
            """
            with recursive tree(id) as (
                select id from registry_nodes where id = ?
                union all
                select registry_nodes.id
                from registry_nodes
                join tree on registry_nodes.parent_id = tree.id
            )
            select id from tree
            """,
            (node_id,),
        ).fetchall()
        return [int(row["id"]) for row in rows]

    def delete_node(self, node_id: int) -> list[int]:
        deleted_ids = self.list_descendant_ids(node_id)
        if not deleted_ids:
            raise KeyError(f"registry node {node_id} not found")
        self.connection.execute("delete from registry_nodes where id = ?", (node_id,))
        self.connection.commit()
        return deleted_ids

    def create_watched_file(
        self,
        node_id: int,
        file_path: str,
        label: str,
        file_kind: str,
        baseline_hash: str,
        last_hash: str,
        status: str = "ok",
    ) -> WatchedFile:
        timestamp = _now()
        cursor = self.connection.execute(
            """
            insert into watched_files
            (node_id, file_path, label, file_kind, baseline_hash, last_hash, status, created_at, updated_at)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (node_id, file_path, label, file_kind, baseline_hash, last_hash, status, timestamp, timestamp),
        )
        self.connection.commit()
        return self.get_watched_file(int(cursor.lastrowid))

    def get_watched_file(self, watched_file_id: int) -> WatchedFile:
        row = self.connection.execute(
            "select * from watched_files where id = ?",
            (watched_file_id,),
        ).fetchone()
        if row is None:
            raise KeyError(f"watched file {watched_file_id} not found")
        return _watched_file_from_row(row)

    def delete_watched_file(self, watched_file_id: int) -> WatchedFile:
        watched = self.get_watched_file(watched_file_id)
        self.connection.execute("delete from watched_files where id = ?", (watched_file_id,))
        self.connection.commit()
        return watched

    def list_watched_files(self, node_id: int) -> list[WatchedFile]:
        rows = self.connection.execute(
            "select * from watched_files where node_id = ? order by label",
            (node_id,),
        ).fetchall()
        return [_watched_file_from_row(row) for row in rows]

    def list_watched_files_for_tree(self, node_id: int) -> list[WatchedFile]:
        node_ids = self.list_descendant_ids(node_id)
        if not node_ids:
            return []
        placeholders = ", ".join("?" for _ in node_ids)
        rows = self.connection.execute(
            f"select * from watched_files where node_id in ({placeholders}) order by node_id, label",
            node_ids,
        ).fetchall()
        return [_watched_file_from_row(row) for row in rows]

    def update_watched_file_status(
        self,
        watched_file_id: int,
        last_hash: str,
        status: str,
        last_error: str = "",
        file_path: str | None = None,
    ) -> WatchedFile:
        timestamp = _now()
        self.connection.execute(
            """
            update watched_files
            set file_path = coalesce(?, file_path),
                last_hash = ?,
                status = ?,
                last_error = ?,
                last_checked_at = ?,
                updated_at = ?
            where id = ?
            """,
            (file_path, last_hash, status, last_error, timestamp, timestamp, watched_file_id),
        )
        self.connection.commit()
        return self.get_watched_file(watched_file_id)

    def create_file_version(
        self,
        watched_file_id: int,
        version_number: int,
        snapshot_path: str,
        sha256: str,
        size_bytes: int,
        change_count: int = 0,
    ) -> FileVersion:
        timestamp = _now()
        cursor = self.connection.execute(
            """
            insert into file_versions
            (watched_file_id, version_number, snapshot_path, sha256, size_bytes, change_count, created_at)
            values (?, ?, ?, ?, ?, ?, ?)
            """,
            (watched_file_id, version_number, snapshot_path, sha256, size_bytes, change_count, timestamp),
        )
        self.connection.commit()
        return self.get_file_version(int(cursor.lastrowid))

    def get_file_version(self, version_id: int) -> FileVersion:
        row = self.connection.execute(
            "select * from file_versions where id = ?",
            (version_id,),
        ).fetchone()
        if row is None:
            raise KeyError(f"file version {version_id} not found")
        return _file_version_from_row(row)

    def list_file_versions(self, watched_file_id: int) -> list[FileVersion]:
        rows = self.connection.execute(
            "select * from file_versions where watched_file_id = ? order by version_number",
            (watched_file_id,),
        ).fetchall()
        return [_file_version_from_row(row) for row in rows]

    def get_latest_file_version(self, watched_file_id: int) -> FileVersion:
        row = self.connection.execute(
            """
            select * from file_versions
            where watched_file_id = ?
            order by version_number desc
            limit 1
            """,
            (watched_file_id,),
        ).fetchone()
        if row is None:
            raise KeyError(f"file version for watched file {watched_file_id} not found")
        return _file_version_from_row(row)

    def create_event(
        self,
        node_id: int,
        watched_file_id: int,
        event_type: str,
        change_count: int = 0,
        report_path: str = "",
        message: str = "",
        risk_count: int = 0,
        risk_summary: str = "",
    ) -> MonitorEvent:
        timestamp = _now()
        cursor = self.connection.execute(
            """
            insert into monitor_events
            (node_id, watched_file_id, event_type, change_count, report_path, message, risk_count, risk_summary, created_at)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (node_id, watched_file_id, event_type, change_count, report_path, message, risk_count, risk_summary, timestamp),
        )
        self.connection.commit()
        return self.get_event(int(cursor.lastrowid))

    def get_event(self, event_id: int) -> MonitorEvent:
        row = self.connection.execute(
            "select * from monitor_events where id = ?",
            (event_id,),
        ).fetchone()
        if row is None:
            raise KeyError(f"monitor event {event_id} not found")
        return _event_from_row(row)

    def list_events(self, node_id: int) -> list[MonitorEvent]:
        rows = self.connection.execute(
            "select * from monitor_events where node_id = ? order by id desc",
            (node_id,),
        ).fetchall()
        return [_event_from_row(row) for row in rows]

    def list_events_for_tree(self, node_id: int) -> list[MonitorEvent]:
        node_ids = self.list_descendant_ids(node_id)
        if not node_ids:
            return []
        placeholders = ", ".join("?" for _ in node_ids)
        rows = self.connection.execute(
            f"select * from monitor_events where node_id in ({placeholders}) order by id desc",
            node_ids,
        ).fetchall()
        return [_event_from_row(row) for row in rows]

    def get_last_event_for_watched_file(self, watched_file_id: int) -> MonitorEvent | None:
        row = self.connection.execute(
            """
            select * from monitor_events
            where watched_file_id = ?
            order by id desc
            limit 1
            """,
            (watched_file_id,),
        ).fetchone()
        return None if row is None else _event_from_row(row)

    def get_comparison_review(self, watched_file_id: int, from_version: int, to_version: int) -> ComparisonReview:
        row = self.connection.execute(
            """
            select * from comparison_reviews
            where watched_file_id = ? and from_version = ? and to_version = ?
            """,
            (watched_file_id, from_version, to_version),
        ).fetchone()
        if row is None:
            timestamp = _now()
            return ComparisonReview(
                id=0,
                watched_file_id=watched_file_id,
                from_version=from_version,
                to_version=to_version,
                status="new",
                reviewer="",
                comment="",
                reviewed_at="",
                updated_at=timestamp,
            )
        return _comparison_review_from_row(row)

    def upsert_comparison_review(
        self,
        watched_file_id: int,
        from_version: int,
        to_version: int,
        status: str,
        reviewer: str = "",
        comment: str = "",
    ) -> ComparisonReview:
        timestamp = _now()
        reviewed_at = timestamp if status != "new" else ""
        self.connection.execute(
            """
            insert into comparison_reviews
            (watched_file_id, from_version, to_version, status, reviewer, comment, reviewed_at, updated_at)
            values (?, ?, ?, ?, ?, ?, ?, ?)
            on conflict(watched_file_id, from_version, to_version) do update set
                status = excluded.status,
                reviewer = excluded.reviewer,
                comment = excluded.comment,
                reviewed_at = excluded.reviewed_at,
                updated_at = excluded.updated_at
            """,
            (watched_file_id, from_version, to_version, status, reviewer, comment, reviewed_at, timestamp),
        )
        self.connection.commit()
        return self.get_comparison_review(watched_file_id, from_version, to_version)

    def node_path(self, node_id: int) -> str:
        names: list[str] = []
        current: int | None = node_id
        while current is not None:
            node = self.get_node(current)
            names.append(node.name)
            current = node.parent_id
        return " / ".join(reversed(names))

    def snapshot_root(self) -> Path:
        if self.db_path is None:
            return Path.cwd() / "registry_snapshots"
        return self.db_path.parent / "registry_snapshots"

    def backup_database(self, target_dir: Path | str | None = None) -> Path:
        if self.db_path is None:
            raise RuntimeError("database backup requires a file-backed SQLite database")
        self.connection.commit()
        backup_dir = Path(target_dir) if target_dir is not None else self.db_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        backup_path = backup_dir / f"registry-backup-{timestamp}.sqlite3"
        with sqlite3.connect(backup_path) as target:
            self.connection.backup(target)
        return backup_path

    def export_registry_excel(self, output_path: Path | str) -> Path:
        export_path = Path(output_path)
        export_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        objects_sheet = workbook.active
        objects_sheet.title = "Объекты"
        files_sheet = workbook.create_sheet("Файлы")
        compare_sheet = workbook.create_sheet("Сверки")
        history_sheet = workbook.create_sheet("История")

        _write_rows(
            objects_sheet,
            [
                ["ID", "Родитель", "Путь объекта", "Название", "Код", "Ответственный", "Тип", "Описание", "Создан", "Обновлен"],
                *self._object_export_rows(),
            ],
        )
        _write_rows(
            files_sheet,
            [
                ["ID", "Путь объекта", "Объект", "Файл", "Тип", "Статус", "Текущий путь", "Версий", "Последняя проверка", "Обновлен"],
                *self._file_export_rows(),
            ],
        )
        _write_rows(
            compare_sheet,
            [
                [
                    "ID",
                    "Файл ID",
                    "Путь объекта",
                    "Файл",
                    "Версии",
                    "Файл до",
                    "Файл после",
                    "Изменений",
                    "Загружено",
                    "Согласование",
                    "Проверил",
                    "Комментарий",
                ],
                *self._comparison_export_rows(),
            ],
        )
        _write_rows(
            history_sheet,
            [
                ["ID", "Путь объекта", "Файл", "Тип события", "Статус", "Изменений", "Сообщение", "Создано"],
                *self._history_export_rows(),
            ],
        )
        for sheet in workbook.worksheets:
            _style_export_sheet(sheet)
        workbook.save(export_path)
        return export_path

    def _object_export_rows(self) -> list[list[object]]:
        rows = self.connection.execute("select * from registry_nodes order by id").fetchall()
        result: list[list[object]] = []
        for row in rows:
            node = _node_from_row(row)
            parent_name = self.get_node(node.parent_id).name if node.parent_id is not None else ""
            result.append(
                [
                    node.id,
                    parent_name,
                    self.node_path(node.id),
                    node.name,
                    node.code,
                    node.responsible,
                    node.node_type,
                    node.description,
                    node.created_at,
                    node.updated_at,
                ]
            )
        return result

    def _file_export_rows(self) -> list[list[object]]:
        rows = self.connection.execute("select * from watched_files order by id").fetchall()
        result: list[list[object]] = []
        for row in rows:
            watched = _watched_file_from_row(row)
            node = self.get_node(watched.node_id)
            result.append(
                [
                    watched.id,
                    self.node_path(watched.node_id),
                    node.name,
                    watched.label,
                    watched.file_kind,
                    _registry_status_text(watched.status),
                    watched.file_path,
                    len(self.list_file_versions(watched.id)),
                    watched.last_checked_at or "",
                    watched.updated_at,
                ]
            )
        return result

    def _comparison_export_rows(self) -> list[list[object]]:
        watched_rows = self.connection.execute("select * from watched_files order by id").fetchall()
        result: list[list[object]] = []
        row_number = 1
        for watched_row in watched_rows:
            watched = _watched_file_from_row(watched_row)
            versions = self.list_file_versions(watched.id)
            for previous, current in zip(versions, versions[1:]):
                review = self.get_comparison_review(watched.id, previous.version_number, current.version_number)
                result.append(
                    [
                        row_number,
                        watched.id,
                        self.node_path(watched.node_id),
                        watched.label,
                        f"v{previous.version_number} -> v{current.version_number}",
                        _snapshot_display_name(previous.snapshot_path),
                        _snapshot_display_name(current.snapshot_path),
                        current.change_count,
                        current.created_at,
                        _review_status_text(review.status),
                        review.reviewer,
                        review.comment,
                    ]
                )
                row_number += 1
        return result

    def _history_export_rows(self) -> list[list[object]]:
        rows = self.connection.execute(
            """
            select monitor_events.*, watched_files.label as file_label
            from monitor_events
            left join watched_files on watched_files.id = monitor_events.watched_file_id
            order by monitor_events.id desc
            """
        ).fetchall()
        result: list[list[object]] = []
        for row in rows:
            event = _event_from_row(row)
            result.append(
                [
                    event.id,
                    self.node_path(event.node_id),
                    row["file_label"] or "",
                    event.event_type,
                    _registry_status_text(event.event_type),
                    event.change_count,
                    event.message,
                    event.created_at,
                ]
            )
        return result


def _node_from_row(row: sqlite3.Row) -> RegistryNode:
    return RegistryNode(
        id=int(row["id"]),
        parent_id=row["parent_id"],
        name=row["name"],
        node_type=row["node_type"],
        code=row["code"],
        responsible=row["responsible"],
        description=row["description"],
        monitoring_enabled=bool(row["monitoring_enabled"]),
        monitor_interval_minutes=int(row["monitor_interval_minutes"]),
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        template_key=row["template_key"],
    )


def _file_version_from_row(row: sqlite3.Row) -> FileVersion:
    return FileVersion(
        id=int(row["id"]),
        watched_file_id=int(row["watched_file_id"]),
        version_number=int(row["version_number"]),
        snapshot_path=row["snapshot_path"],
        sha256=row["sha256"],
        size_bytes=int(row["size_bytes"]),
        change_count=int(row["change_count"]),
        created_at=row["created_at"],
    )


def _watched_file_from_row(row: sqlite3.Row) -> WatchedFile:
    return WatchedFile(
        id=int(row["id"]),
        node_id=int(row["node_id"]),
        file_path=row["file_path"],
        label=row["label"],
        file_kind=row["file_kind"],
        baseline_hash=row["baseline_hash"],
        last_hash=row["last_hash"],
        last_checked_at=row["last_checked_at"],
        status=row["status"],
        last_error=row["last_error"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
    )


def _event_from_row(row: sqlite3.Row) -> MonitorEvent:
    return MonitorEvent(
        id=int(row["id"]),
        node_id=int(row["node_id"]),
        watched_file_id=int(row["watched_file_id"]),
        event_type=row["event_type"],
        change_count=int(row["change_count"]),
        report_path=row["report_path"],
        message=row["message"],
        created_at=row["created_at"],
        risk_count=int(row["risk_count"]),
        risk_summary=row["risk_summary"],
    )


def _comparison_review_from_row(row: sqlite3.Row) -> ComparisonReview:
    return ComparisonReview(
        id=int(row["id"]),
        watched_file_id=int(row["watched_file_id"]),
        from_version=int(row["from_version"]),
        to_version=int(row["to_version"]),
        status=row["status"],
        reviewer=row["reviewer"],
        comment=row["comment"],
        reviewed_at=row["reviewed_at"],
        updated_at=row["updated_at"],
    )


def _connection_path(connection: sqlite3.Connection) -> Path | None:
    row = connection.execute("pragma database_list").fetchone()
    if row is None:
        return None
    file_path = row["file"] if isinstance(row, sqlite3.Row) else row[2]
    return Path(file_path) if file_path else None


def _registry_status_text(status: str) -> str:
    return {
        "ok": "OK",
        "baseline": "Базовая версия",
        "changed": "Изменен",
        "missing": "Нет файла",
        "error": "Ошибка",
    }.get(status, status)


def _review_status_text(status: str) -> str:
    return {
        "new": "Новое",
        "checked": "Проверено",
        "approved": "Согласовано",
        "rejected": "Отклонено",
    }.get(status, status)


def _snapshot_display_name(snapshot_path: str) -> str:
    name = Path(snapshot_path).name
    return name[6:] if len(name) > 6 and name.startswith("v") and name[1:5].isdigit() and name[5] == "_" else name


def _write_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_export_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="EFF6FF")
    header_font = Font(bold=True, color="1F3A5F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 70))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
